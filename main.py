#!/usr/bin/env python

from PyQt4 import QtGui # Import the PyQt4 module we'll need
from PyQt4.QtCore import QSettings, QSize, QPoint, QCoreApplication, QThread, SIGNAL # pylint: disable-msg=E0611
from pathlib import Path
from dotenv import load_dotenv
from fields import sections, topSection, bottomSection, possibleSize
from fields import startSections, startSectionsSize, endSections, partSection, partSectionByModel, costSummary, boatLength
import openpyxl
import pickle
import sys # We need sys so that we can pass argv to QApplication
import os
import re
import MainWindow  # This file holds our MainWindow and all design related things



r"""
Notes:
patch of the PyInstaller/depend/bindepend.py https://github.com/Loran425/pyinstaller/commit/14b6e65642e4b07a4358bab278019a48dedf7460

To design UI: Lib\site-packages\PyQt4\Designer.exe

To rebuild UI: Lib\site-packages\PyQt4\pyuic4 MainWindow.ui  -o MainWindow.py

Developed in C:\\Development\\nrb-pickle-boats :
Scripts\pyinstaller.exe --onefile --windowed --icon options.ico  --name "Boat Folder Pickler" "NRB Pickle Boats FWW.spec" main.py

ToDo's
"""


class MainAppWindow(QtGui.QMainWindow, MainWindow.Ui_MainWindow):

    def __init__(self):
        super().__init__()
        self.setupUi(self)

        # set python environment
        if getattr(sys, 'frozen', False):
            bundle_dir = sys._MEIPASS # pylint: disable=no-member
        else:
            # we are running in a normal Python environment
            bundle_dir = os.path.dirname(os.path.abspath(__file__))

        # load environmental variables
        load_dotenv(dotenv_path = Path(bundle_dir) / ".env")

        # set program icon
        self.setWindowIcon(QtGui.QIcon(os.path.join(bundle_dir, "pickle.ico")))

        # work in INI File Stuff here
        QCoreApplication.setOrganizationName("NRB")
        QCoreApplication.setOrganizationDomain("northriverboats.com")
        QCoreApplication.setApplicationName("Options Boat Pickler")
        self.settings = QSettings()
        

        # set variables
        self.background_thread = None
        self.exit_flag = False
        self.dir = self.settings.value("dir", os.getenv("DIR"))
        self.pickle_name = os.getenv("PICKLE")

        # set ui state
        self.actionCancel.setEnabled(False)
        self.btnCancel.hide()
        self.lePath.setText(self.dir)

        # set slots and signals
        self.actionExit.triggered.connect(self.closeEvent)
        self.actionAbout.triggered.connect(self.doAbout)
        self.btnBrowse.clicked.connect(self.browseEvent)
        self.btnRun.clicked.connect(self.startBackgroundTask)

    def doAbout(self, event):
        about_msg = "NRB Boat Folder Pickler\n©2019 North River Boats\nBy Fred Warren"
        QtGui.QMessageBox.information(self, 'About',
                about_msg, QtGui.QMessageBox.Ok)

    def closeEvent(self, e):
        self._closeEvent(0)

    def _closeEvent(self, e):
        self.exit_flag = True
        self.settings.setValue("dir", self.dir)
        sys.exit(0)

    def browseEvent(self):
        default_dir = self.dir
        my_dir = QtGui.QFileDialog.getExistingDirectory(self, "Open a folder", default_dir, QtGui.QFileDialog.ShowDirsOnly)
        if my_dir != "":
            self.dir = my_dir
            self.update_progressbar(0)
        self.lePath.setText(self.dir)



    def startBackgroundTask(self):
        # hide / disable buttons and menu items as well as saving state
        self.block_actions()
        self.background_thread = background_thread(self.dir)

        self.connect(self.background_thread, SIGNAL('endBackgroundTask()'), self.endBackgroundTask)
        self.connect(self.background_thread, SIGNAL('update_statusbar(QString)'), self.update_statusbar)
        self.connect(self.background_thread, SIGNAL('update_label(QString)'), self.update_label)
        self.connect(self.background_thread, SIGNAL('update_progressbar(int)'), self.update_progressbar)
        self.btnCancel.clicked.connect(self.doAbort)    
        # Thread will self-terminate or be stopped via update_abort
        self.background_thread.start()

    def endBackgroundTask(self):
        self.unblock_actions()
        self.statusbar.showMessage("")

    def doAbort(self):
        self.background_thread.running = False
        self.unblock_actions

    def block_actions(self):        
        self.btnRun.setEnabled(False)
        self.btnCancel.setEnabled(True)
        self.btnRun.hide()
        self.btnCancel.show()
        self.actionRun.setEnabled(False)
        self.actionCancel.setEnabled(True)
        self.lePath.setReadOnly(True)
    
    def unblock_actions(self):
        self.btnRun.setEnabled(True)
        self.btnCancel.setEnabled(False)
        self.btnRun.show()
        self.btnCancel.hide()
        self.actionRun.setEnabled(True)
        self.actionCancel.setEnabled(False)
        self.lePath.setReadOnly(False)
        self.lblFile.setText("")

    def update_statusbar(self, message):
        self.statusbar.showMessage(message)

    def update_progressbar(self, num):
        self.progressBar.setValue(num)
        
    def update_label(self, label):
        self.lblFile.setText(label)


        
class background_thread(QThread):
    def __init__(self, dir):
        QThread.__init__(self)
        self.dir = dir
        self.wb = None
        self.ws = None
        self.starts = []
        self.ends = []
        self.boatSizes = []
        self.data = {}
        self.part = {}

    def __del__(self):
        self.wait()

    def build_files_list(self, dir):
        self.emit(SIGNAL('update_statusbar(qString)'), "Finding files...")
        files = []
        for path in Path(dir).glob("[!~$]*.xlsx"):
            if not self.running:
                break
            files.append(path)
        self.emit(SIGNAL('update_statusbar(qString)'), "Found {} files to process".format(len(files)))
        return files

    def open_sheet(self, file):
        self.wb = openpyxl.load_workbook(file, data_only = True)
        self.ws = self.wb.active

    #### IDENTIFY PORTIONS OF SHEET TO PROCESS #############
    def find_starts(self):
        # find where sections start
        self.starts = []
        for row in self.ws.iter_cols(min_col=8, max_col=8):
            for cell in row:
                if cell.value == "QTY.":
                    self.starts.append(cell.row)

    def find_ends(self):
        # find where sections end
        self.ends = []
        for row in self.ws.iter_cols(min_col=10, max_col=10):
            for cell in row:
                if cell.value == "SUBTOTAL":
                    self.ends.append(cell.row)
    
    def find_boat_sizes(self):
        self.boatSizes = []
        for col in self.ws.iter_rows(min_row=1, max_row=1):
            for cell in col:
                if  str(cell.value).isdigit():
                    self.boatSizes.append(cell.value)
    
    #### PROCESS NON-SECTION PORTIONS OF THE SHEET #############    
    def process_top_section(self):
        # Process top static section of sheet
        for name, column, row, default in topSection:
            value = self.ws.cell(column = column, row = row).value
            if value is None:
                value = default
            self.data[name] = value

    def process_cost_summary_by_boat_size(self):
        #Process cost summary
        for i, boatSize in enumerate(self.boatSizes):
            for name, column, row, default in costSummary:
                value = self.ws.cell(column = column + (i * 4), row = row).value
                if value is None:
                    value = default
                self.data[str(boatSize) + name] = value

    #### PROCESS SECTION PORTIONS OF THE SHEET SUPPORTING FUNCTIONS #############
    def process_inner_section_top(self, index, offset, section):
        for name, column, row, default in startSections:
            value = self.ws.cell(column = column, row = row + offset).value
            if value is None:
                value = default
            # print(section + name, value)
            self.data[section + name] = value

    def process_inner_section_top_by_boat_size(self, index, offset, boatSize, section):
        for name, column, row, default in startSectionsSize:
            value = self.ws.cell(column = column + (index * 4), row = row + offset).value
            if value is None:
                value = default
            # print(section + " " + str(boatSize) + name, value)
            self.data[section + " " + str(boatSize) + name] = value

    def process_inner_section_bottom_by_boat_size(self, index, offset, boatSize, section):
        for name, column, row, default in endSections:
            value = self.ws.cell(column = column + (index * 4), row = row + offset).value
            if value is None:
                value = default
            # print(section + " " + str(boatSize) + name, value)
            self.data[section + " " + str(boatSize) + name] = value


    def process_part_by_boat_size(self, index, offset, boatSize, section):
        for name, column, row, default in partSectionByModel:
            value = self.ws.cell(column = column + (index * 4), row = row + offset).value
            if value is None:
                value = default
            self.part[str(boatSize) + name] = value

    def process_part_by_non_boat_size(self, offset, section):
        for name, column, row, default in partSection:
            value = self.ws.cell(column = column, row = row + offset).value
            if value is None:
                value = default
            self.part[name] = value


    def process_section_by_boat_sizes(self, offset, section, process_by_size_function):
        for index, boatSize in enumerate(self.boatSizes):
            process_by_size_function(index, offset, boatSize, section)

    def process_parts(self, index, section):
        for offset in range(self.starts[index], self.ends[index]-1):
            self.part = {}
            if self.ws.cell(column = 1, row = 1 + offset).value is not None:
                self.process_part_by_non_boat_size(offset, section)
                self.process_section_by_boat_sizes(offset, section, self.process_part_by_boat_size)
                self.data[section + " PARTS"].append(self.part)

                
    #### PROCESS SECTION PORTIONS OF THE SHEET #############
    def process_section_top(self):
        # Process top non-parts portion of sections not by boat size
        for index, section in enumerate(sections):
            offset = self.starts[index]
            self.process_inner_section_top(index, offset, section)

    def process_section_top_by_boat_size(self):
		# Process top non-parts portion of sections by boat size
        for i, section in enumerate(sections):
            offset = self.starts[i]
            self.process_section_by_boat_sizes(offset, section, self.process_inner_section_top_by_boat_size)
    
    def process_section_parts(self):
        # Process parts portion of sections both by non boat size and by boat size must be combined
        for index, section in enumerate(sections):
            self.data[section + " PARTS"] = []
            self.process_parts(index, section)
            # print(section + " PARTS", self.data[section + " PARTS"])
    
    def process_section_bottom(self):
        # Process bottom non-parts portion of sections
        for i, section in enumerate(sections):
            offset = self.ends[i]
            self.process_section_by_boat_sizes(offset, section, self.process_inner_section_bottom_by_boat_size)


    #### PROCESS FULL SHEET ##########################
    def process_sheet(self, file):
        self.open_sheet(file)

        # find different sections on the sheet
        self.find_starts()
        self.find_ends()
        self.find_boat_sizes()

        self.data = {}

        # keys not derived from fields.py
        self.data["FILE"] = file.name
        self.data["FULLPATH"] = file.resolve()
        self.data['BOAT SIZES'] = self.boatSizes

        # process top non-section band
        self.process_top_section() # only non-boat-size
        self.process_cost_summary_by_boat_size() # only by-boat-size
        
        # process section bands
        self.process_section_top() # only non-boat-size
        self.process_section_top_by_boat_size() # only by-boat-size
        self.process_section_parts() # both non-boat-size and by-boat-size
        self.process_section_bottom() # only non-boat-size

        return [str(self.data["BOAT MODEL"]), self.data]

    def run(self):
        self.running = True
        self.emit(SIGNAL('update_progressbar(int)'), 0)
        files = self.build_files_list(self.dir)
        options = {}
        total_files = len(files)
        current_count = 0

        if not self.running:
            self.emit(SIGNAL('endBackgroundTask()'))
            return

        for file in files:
            if not self.running:
                break

            option, data = self.process_sheet(file)
            options[option] = data

            current_count += 1
            self.emit(SIGNAL('update_progressbar(int)'), int(float(current_count) / total_files * 100))
            self.emit(SIGNAL('update_label(QString)'), str(file))
            self.emit(SIGNAL('update_statusbar(QString)'), 'Pickeling %d of %d' % (current_count, total_files))

        # if we get to this point, pickle the results....
        file_name = os.path.join(self.dir, os.path.split(self.dir)[1].lower() + ".pickle")
        pickle.dump(options, open(file_name, 'wb'))
        self.emit(SIGNAL('endBackgroundTask()'))


def main():
    app = QtGui.QApplication(sys.argv)  # A new instance of QApplication
    form = MainAppWindow()              # We set the form to be our Main App Wehdiw (design)
    form.show()                         # Show the form
    app.exec_()                         # and execute the app


if __name__ == '__main__':              # if we're running file directly and not importing it
    main()                              # run the main function